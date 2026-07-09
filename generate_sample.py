"""
Generate sample Excel with merged header cells (3 rows) and null rows.
Simulates real-world messy Excel data that Power Query needs to clean.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from copy import copy

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Merged 3-row header + some null rows
# ============================================================
ws1 = wb.active
ws1.title = "Sales_Report"

# --- Row 1: Top-level category (merged across columns) ---
ws1.merge_cells("A1:B1")
ws1["A1"] = "Product Info"
ws1.merge_cells("C1:E1")
ws1["C1"] = "Q1 Sales"
ws1.merge_cells("F1:H1")
ws1["F1"] = "Q2 Sales"
# Column I: header left blank across all 3 rows → exercises the
# "all-empty header" fallback (renamed to Column_I).
# Column J: junk header + junk data → exercises the case where a
# non-main column keeps an otherwise-fully-null row alive.
ws1["J1"] = "asdasd"
ws1["J2"] = "aaa"
ws1["J3"] = "a"

# --- Row 2: Sub-category ---
ws1["A2"] = "ID"
ws1["B2"] = "Name"
ws1.merge_cells("C2:D2")
ws1["C2"] = "Revenue"
ws1["E2"] = "Units"
ws1.merge_cells("F2:G2")
ws1["F2"] = "Revenue"
ws1["H2"] = "Units"

# --- Row 3: Detail header ---
ws1["A3"] = "ID"
ws1["B3"] = "Name"
ws1["C3"] = "Jan"
ws1["D3"] = "Feb"
ws1["E3"] = "Mar"
ws1["F3"] = "Apr"
ws1["G3"] = "May"
ws1["H3"] = "Jun"

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
for row in range(1, 4):
    for col in range(1, 11):
        cell = ws1.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

# --- Data rows (starting row 4) ---
# Column I mirrors column H (junk duplicate, header left blank).
# Column J has real junk data on row 6 so that row survives null-row removal.
data = [
    ["P001", "Widget A",   1000, 1200, 1100, 1300, 1400, 1250, 1250, 1250],
    ["P002", "Gadget B",    800,  900,  850,  950, 1000,  920,  920,  920],
    [None,   None,         None, None, None, None, None, None, "sss", "aaa"],  # kept alive by col J junk
    ["P003", "Doohickey C", 500,  600,  550,  700,  650,  680,  680,  680],
    [None,   None,         None, None, None, None, None, None, None, None],  # null row
    [None,   None,         None, None, None, None, None, None, None, None],  # null row
    ["P004", "Thingamajig", 2000, 2100, 2200, 2300, 2400, 2150, 2150, 2150],
    ["P005", "Whatchamacallit", 300, 350, 400, 450, 500, 420, 420, 420],
    [None,   None,         None, None, None, None, None, None, None, None],  # null row
    ["P006", "Contraption", 1500, 1600, 1550, 1650, 1700, 1580, 1580, 1580],
]

for i, row_data in enumerate(data):
    row_num = i + 4
    for j, val in enumerate(row_data):
        cell = ws1.cell(row=row_num, column=j + 1)
        cell.value = val
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# ============================================================
# Sheet 2: Another variant — merged headers with different structure
# ============================================================
ws2 = wb.create_sheet("Employee_Data")

ws2.merge_cells("A1:A3")
ws2["A1"] = "Emp ID"
ws2.merge_cells("B1:C1")
ws2["B1"] = "Name"
ws2["B2"] = "First"
ws2["C2"] = "Last"
ws2.merge_cells("D1:F1")
ws2["D1"] = "Contact"
ws2["D2"] = "Email"
ws2["E2"] = "Phone"
ws2["F2"] = "Address"
# Row 3 is merged into A1:A3, B1:C1, D1:F1 — so only A3, B3, C3, D3, E3, F3 exist as merged
# Actually let me redo this more carefully
# Clear and redo
ws2.delete_rows(1, 3)

ws2.merge_cells("A1:A3")
ws2["A1"] = "Emp ID"
ws2.merge_cells("B1:C1")
ws2["B1"] = "Name"
ws2["B2"] = "First"
ws2["C2"] = "Last"
ws2.merge_cells("D1:F1")
ws2["D1"] = "Contact"
ws2["D2"] = "Email"
ws2["E2"] = "Phone"
ws2["F2"] = "Address"

for row in range(1, 4):
    for col in range(1, 7):
        cell = ws2.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

emp_data = [
    ["E001", "Alice", "Smith",   "alice@co.com", "555-0101", "123 Main St"],
    ["E002", "Bob",   "Jones",   "bob@co.com",   "555-0102", "456 Oak Ave"],
    [None,   None,    None,      None,           None,       None],
    ["E003", "Carol", "Williams","carol@co.com", "555-0103", "789 Pine Rd"],
    ["E004", "Dan",   "Brown",   "dan@co.com",   "555-0104", "321 Elm St"],
    [None,   None,    None,      None,           None,       None],
]

for i, row_data in enumerate(emp_data):
    row_num = i + 4
    for j, val in enumerate(row_data):
        cell = ws2.cell(row=row_num, column=j + 1)
        cell.value = val
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# ============================================================
# Sheet 3: Clean reference data (no merged headers, no nulls)
# ============================================================
ws3 = wb.create_sheet("Reference")
ws3["A1"] = "Code"
ws3["B1"] = "Description"
ws3["C1"] = "Category"
for col in range(1, 4):
    cell = ws3.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

ref_data = [
    ["P001", "Widget A", "Type 1"],
    ["P002", "Gadget B", "Type 2"],
    ["P003", "Doohickey C", "Type 1"],
    ["P004", "Thingamajig", "Type 3"],
    ["P005", "Whatchamacallit", "Type 2"],
    ["P006", "Contraption", "Type 3"],
]
for i, row_data in enumerate(ref_data):
    row_num = i + 2
    for j, val in enumerate(row_data):
        cell = ws3.cell(row=row_num, column=j + 1)
        cell.value = val
        cell.border = thin_border

# ============================================================
# Sheet 4: Edge case — some columns have null headers across all 3 rows
# ============================================================
ws4 = wb.create_sheet("Edge_NullHeaders")

# Column A: completely null across all 3 header rows (no left neighbor to fill from)
# Column B: has "Item"
# Column C: null all 3 rows but FillRight will fill from B → so it gets a value
# Column D: has "Price"
ws4["A1"] = None
ws4["B1"] = "Item"
ws4["C1"] = None
ws4["D1"] = "Price"

ws4["A2"] = None
ws4["B2"] = "Item"
ws4["C2"] = None
ws4["D2"] = "Price"

ws4["A3"] = None
ws4["B3"] = "Item"
ws4["C3"] = None
ws4["D3"] = "Price"

# So column A header = null all 3 rows → fallback "Column1"
# Column B = "Item", C = "Item" (filled right), D = "Price"

for row in range(1, 4):
    for col in range(1, 5):
        cell = ws4.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

edge_data = [
    ["x",  "Apple",  "red",    10],
    ["y",  "Banana", "yellow", 20],
    [None, None,     None,     None],  # null row
    ["z",  "Cherry", "red",    30],
]
for i, row_data in enumerate(edge_data):
    row_num = i + 4
    for j, val in enumerate(row_data):
        cell = ws4.cell(row=row_num, column=j + 1)
        cell.value = val
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# ============================================================
# Sheet 5: Edge case — duplicate header names after combine
# ============================================================
ws5 = wb.create_sheet("Edge_DupHeaders")

# Two columns both named "Value" across all 3 header rows
ws5["A1"] = "ID"
ws5["B1"] = "Value"
ws5["C1"] = "Value"   # duplicate!
ws5["D1"] = "Value"   # another duplicate!

ws5["A2"] = "ID"
ws5["B2"] = "Value"
ws5["C2"] = "Value"
ws5["D2"] = "Value"

ws5["A3"] = "ID"
ws5["B3"] = "Value"
ws5["C3"] = "Value"
ws5["D3"] = "Value"

for row in range(1, 4):
    for col in range(1, 5):
        cell = ws5.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

dup_data = [
    ["X1", 100, 200, 300],
    [None, None, None, None],  # null row
    ["X2", 400, 500, 600],
]
for i, row_data in enumerate(dup_data):
    row_num = i + 4
    for j, val in enumerate(row_data):
        cell = ws5.cell(row=row_num, column=j + 1)
        cell.value = val
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# ============================================================
# Sheet 6: Edge case — column mixes numbers with non-numeric text
# ============================================================
ws6 = wb.create_sheet("Edge_MixedTypes")

# "Amount" looks numeric at a glance, but row 3 has "N/A" — a
# first-value-only type check would misdetect this as numeric.
ws6["A1"] = "Code"
ws6["B1"] = "Amount"
for col in range(1, 3):
    cell = ws6.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

mixed_data = [
    ["A1", 100],
    ["A2", 200],
    ["A3", "N/A"],
    ["A4", 300],
]
for i, row_data in enumerate(mixed_data):
    row_num = i + 2
    for j, val in enumerate(row_data):
        cell = ws6.cell(row=row_num, column=j + 1)
        cell.value = val
        cell.border = thin_border

# Save
output_path = r"c:\Users\User\Downloads\excel_proj\sample_data.xlsx"
wb.save(output_path)
print(f"Sample Excel saved to: {output_path}")
print("Sheets: Sales_Report (merged 3-row headers + null rows)")
print("        Employee_Data (merged 3-row headers + null rows)")
print("        Reference (clean, no issues)")
print("        Edge_NullHeaders (some columns have null headers)")
print("        Edge_DupHeaders (duplicate header names after combine)")
print("        Edge_MixedTypes (numeric column with a stray text value)")