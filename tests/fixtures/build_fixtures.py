"""
Builds test fixtures for Daily_Update_HeaderList.pq / Daily_Update_Combined.pq:

  - ReportWorkbook.xlsx  — the "report" workbook where the queries live.
                           Has ConfigTable and SelectColumnTable already set
                           up as real Excel Tables.
  - daily_files/Daily_Update_2568.xlsx
  - daily_files/Daily_Update_2569.xlsx
                           — two source files with a "Daily Update" sheet:
                           3 junk/title rows, a 3-row merged header, then
                           data rows. Mimics the real production layout.

Uses .xlsx (not .xlsm) for the source files because openpyxl can't produce
genuine macro-enabled workbooks — Excel would flag a fake .xlsm as a
format/extension mismatch. ConfigTable's FileExtension is set to ".xlsx"
here; that's just a config value, change it back to ".xlsm" for real data.

Run: ../../.venv/Scripts/python.exe build_fixtures.py
"""
import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
DAILY_DIR = os.path.join(HERE, "daily_files")


def add_table(ws, ref, name):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tbl)


def build_report_workbook():
    wb = Workbook()

    cfg = wb.active
    cfg.title = "Config"
    cfg.append(["Setting", "Value", "หมายเหตุ"])
    cfg.append(["FolderPath", DAILY_DIR, "โฟลเดอร์ไฟล์ต้นทาง (ทดสอบ)"])
    cfg.append(["SheetName", "Daily Update", "ชื่อ sheet ที่ดึง"])
    cfg.append(["JunkRows", 3, "แถวขยะบนสุดที่ข้าม"])
    cfg.append(["HeaderRows", 3, "แถว merged header"])
    cfg.append(["Separator", "_", "ตัวคั่นชื่อ header"])
    cfg.append(["FileExtension", ".xlsx", "นามสกุลไฟล์ (ทดสอบ = .xlsx, ของจริง = .xlsm)"])
    add_table(cfg, "A1:C7", "ConfigTable")

    sel = wb.create_sheet("select column")
    sel.append(["ColumnName"])
    for name in ["ข้อมูลพนักงาน_รหัส", "ข้อมูลพนักงาน_จำนวนเงิน", "SourceFile", "ปี"]:
        sel.append([name])
    add_table(sel, "A1:A5", "SelectColumnTable")

    out = os.path.join(HERE, "ReportWorkbook.xlsx")
    wb.save(out)
    print("wrote", out)


def build_daily_file(path, id_values, amount_values):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Update"

    # Rows 1-3: junk/title rows (skipped via JunkRows=3)
    ws.append(["รายงานประจำวัน — สร้างโดยระบบอัตโนมัติ"])
    ws.append(["แผนก: บัญชี"])
    ws.append([])

    # Rows 4-6: real 3-row merged header
    #   row4: section title merged across all 3 columns
    #   row5: actual field names
    #   row6: blank (filled down from row5 by fnCleanMergedHeaders)
    ws.append(["ข้อมูลพนักงาน", None, None])
    ws.append(["รหัส", "ชื่อ", "จำนวนเงิน"])
    ws.append([None, None, None])
    ws.merge_cells("A4:C4")

    # Data rows (row 7+)
    for id_val, name_val, amount_val in zip(id_values, ["สมชาย", "สมหญิง", "วิชัย"], amount_values):
        ws.append([id_val, name_val, amount_val])

    wb.save(path)
    print("wrote", path)


if __name__ == "__main__":
    os.makedirs(DAILY_DIR, exist_ok=True)
    build_report_workbook()
    build_daily_file(
        os.path.join(DAILY_DIR, "Daily_Update_2568.xlsx"),
        id_values=["E001", "E002", "E003"],
        amount_values=[1000, 2000, 3000],
    )
    build_daily_file(
        os.path.join(DAILY_DIR, "Daily_Update_2569.xlsx"),
        id_values=["E101", "E102", "E103"],
        amount_values=[1500, 2500, 3500],
    )
