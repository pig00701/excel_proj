"""
Test harness: Validates the Power Query M logic using Python.
Simulates fnCleanMergedHeaders algorithm on the sample Excel data.
"""
import openpyxl
from pathlib import Path
from typing import Any, Optional

# ============================================================
# Replicate the M logic in Python
# ============================================================

def fill_right(lst: list, blank_flags: list[bool]) -> list:
    """
    Fill nulls left-to-right (simulates merged cells in a row).
    blank_flags marks columns that are blank across every header row (a true
    gap, not part of any merge) — those positions always stay None: they
    don't receive a fill from the left, and they don't pass anything to
    their right neighbor either. Without this, a genuinely blank column
    next to a real one gets silently absorbed into that real column's name.
    """
    result = []
    last_value = None
    for current, is_blank in zip(lst, blank_flags):
        if is_blank:
            result.append(None)
            last_value = None
        elif current is not None:
            result.append(current)
            last_value = current
        elif last_value is not None:
            result.append(last_value)
        else:
            result.append(None)
    return result


def fill_down(rows: list[list]) -> list[list]:
    """Fill nulls top-to-bottom within header section."""
    if not rows:
        return rows
    num_cols = len(rows[0])
    for col in range(num_cols):
        last_val = None
        for row in rows:
            if row[col] is not None:
                last_val = row[col]
            elif last_val is not None:
                row[col] = last_val
    return rows


def combine_headers(header_rows: list[list], sep: str = "_") -> list[str]:
    """Combine multi-row headers into single row, deduplicating consecutive identical parts."""
    num_cols = len(header_rows[0])
    result = []
    for col_idx in range(num_cols):
        parts = []
        for row in header_rows:
            val = row[col_idx]
            parts.append("" if val is None else str(val))
        # Remove consecutive duplicates
        deduped = []
        for p in parts:
            if len(deduped) > 0 and p == deduped[-1]:
                continue
            deduped.append(p)
        # Remove empty parts
        non_empty = [p for p in deduped if p != ""]
        result.append(sep.join(non_empty))
    return result


def fix_empty_and_duplicates(names: list[str]) -> list[str]:
    """
    Handle empty/null headers and duplicate names.
    - Empty → "Column_A", "Column_B", ... (Excel column letter, true position)
    - Duplicates → suffix "_2", "_3", ...
    Mirrors the M logic in fnCleanMergedHeaders Step 5.
    """
    # Fallback for empty names
    with_fallback = []
    for idx, name in enumerate(names, start=1):
        if name is None or name == "":
            with_fallback.append(f"Column_{col_index_to_letter(idx)}")
        else:
            with_fallback.append(name)

    # Deduplicate
    seen = {}
    result = []
    for name in with_fallback:
        count = seen.get(name, 0)
        if count == 0:
            result.append(name)
        else:
            result.append(f"{name}_{count + 1}")
        seen[name] = count + 1
    return result


def is_all_null(row: list) -> bool:
    """Check if all values in a row are None."""
    return all(v is None for v in row)


def is_convertible_to_number(val: Any) -> bool:
    """Mirrors M's Number.From(...) success/failure for our sample data."""
    if isinstance(val, (int, float)):
        return True
    try:
        float(str(val))
        return True
    except (TypeError, ValueError):
        return False


def infer_column_type(values: list, mode: str = "all") -> str:
    """
    Mirrors fnCleanMergedHeaders Step 9 (type auto-detection).
    mode="first": old buggy behavior — only checks the first non-null value.
    mode="all":   fixed behavior — every non-null value must convert.
    """
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "text"
    if mode == "first":
        return "number" if is_convertible_to_number(non_null[0]) else "text"
    return "number" if all(is_convertible_to_number(v) for v in non_null) else "text"


def col_index_to_letter(idx: int) -> str:
    """1-based index → Excel column letter (1→A, 2→B, ..., 26→Z, 27→AA)."""
    result = ""
    n = idx
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def select_columns(col_names: list[str], data: list[list], keep: list[str]) -> tuple[list[str], list[list]]:
    """
    Mirrors M's Table.SelectColumns(..., keep, MissingField.Ignore):
    result columns follow the ORDER of `keep`, names not present are skipped.
    """
    idxs = [col_names.index(k) for k in keep if k in col_names]
    return [col_names[i] for i in idxs], [[row[i] for i in idxs] for row in data]


def extract_year(filename: str) -> Optional[int]:
    """Mirrors ExtractYear in Daily_Update_Combined.pq: last 4 chars of base name."""
    base = filename.rsplit(".", 1)[0]
    last4 = base[-4:]
    return int(last4) if last4.isdigit() else None


def year_passes_filter(year: Optional[int], year_from: Optional[int], year_to: Optional[int]) -> bool:
    """
    Mirrors the YearFilteredFiles step in Daily_Update_Combined.pq:
    no bounds = everything passes; a null year is KEPT even when a filter
    is set (off-convention file, but never silently dropped).
    """
    if year_from is None and year_to is None:
        return True
    if year is None:
        return True
    return (year_from is None or year >= year_from) and (year_to is None or year <= year_to)


def archive_file_passes(year: Optional[int], current_year: int,
                        year_from: Optional[int]) -> bool:
    """
    Mirrors the ArchiveFiles filter in Daily_Update_Archive.pq:
    named-year files strictly below CurrentYear (>= YearFrom when set);
    null-year files never archive — they stay on the live side.
    """
    if year is None:
        return False
    return year < current_year and (year_from is None or year >= year_from)


def current_file_passes(year: Optional[int], current_year: int,
                        year_to: Optional[int]) -> bool:
    """
    Mirrors the split-mode YearFilteredFiles filter in
    Daily_Update_Combined.pq: CurrentYear onward (<= YearTo when set),
    plus every null-year file.
    """
    if year is None:
        return True
    return year >= current_year and (year_to is None or year <= year_to)


def clean_sheet(raw_data: list[list], header_rows: int, sep: str = "_",
                columns_to_keep: Optional[list[str]] = None) -> tuple[list[str], list[list]]:
    """
    Clean a sheet with merged multi-row headers and null rows.
    Returns (column_names, data_rows).

    columns_to_keep mirrors fnCleanMergedHeaders' ColumnsToKeep: when given,
    columns are trimmed right after names are built — BEFORE null-row
    removal — so a row that is null in every KEPT column is dropped even if
    it had values in dropped columns.

    IMPORTANT: fallback naming ("Column_A", "Column_B", ...) happens AFTER
    FillRight/FillDown, not before. Injecting the placeholder text into the
    header cells first (the original approach) gives FillRight something to
    leak into a genuinely-empty cell of a NEIGHBORING real column, corrupting
    that column's combined name with an unrelated "_Column_X" suffix.
    """
    # Step 1: Extract header section
    header_section = [list(row) for row in raw_data[:header_rows]]

    # Identify columns blank across ALL header rows (true gaps, not merges)
    # BEFORE any filling, so FillRight can protect them.
    num_cols = len(header_section[0])
    is_blank_column = [
        all(row[col] is None or str(row[col]).strip() == "" for row in header_section)
        for col in range(num_cols)
    ]

    # Step 2: Fill RIGHT within each header row (raw values only)
    header_section = [fill_right(row, is_blank_column) for row in header_section]

    # Step 3: Fill DOWN within header section
    header_section = fill_down(header_section)

    # Step 4: Combine headers
    raw_names = combine_headers(header_section, sep)

    # Step 5: Handle empty (fallback "Column_A", "Column_B", ...) + duplicates
    col_names = fix_empty_and_duplicates(raw_names)

    # Step 6: Get data section
    data_section = [list(row) for row in raw_data[header_rows:]]

    # Step 6.5: Trim to requested columns (before the null-row scan,
    # mirroring the M SelectedData step)
    if columns_to_keep is not None:
        col_names, data_section = select_columns(col_names, data_section, columns_to_keep)

    # Step 7: Remove fully-null rows
    data_section = [row for row in data_section if not is_all_null(row)]

    return col_names, data_section


# ============================================================
# Read Excel and test
# ============================================================

def read_sheet_raw(filepath: str, sheet_name: str) -> list[list]:
    """Read a sheet as raw list-of-lists (no header processing)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        rows.append([cell.value for cell in row])
    wb.close()
    return rows


def run_tests():
    filepath = str(Path(__file__).resolve().parent.parent / "sample_data.xlsx")
    all_pass = True

    # ========================
    # Test 1: Sales_Report
    # ========================
    print("=" * 60)
    print("TEST: Sales_Report (3-row merged headers + null rows)")
    print("=" * 60)

    raw = read_sheet_raw(filepath, "Sales_Report")
    print(f"  Raw rows: {len(raw)} (including headers)")

    col_names, data = clean_sheet(raw, header_rows=3, sep="_")

    print(f"  Column names: {col_names}")
    print(f"  Data rows: {len(data)}")

    # Assertions
    tests = []

    # Test 1.1: Column count = 10 (I renamed to Column_I, J kept as junk)
    tests.append(("Column count = 10", len(col_names) == 10))

    # Test 1.2: Row count = 8 (fully-null rows removed; the two junk rows
    # with values only in Column_I / asdasd_aaa_a survive)
    tests.append(("Row count = 8", len(data) == 8))

    # Test 1.3: No null rows
    null_rows = [i for i, row in enumerate(data) if is_all_null(row)]
    tests.append(("No null rows", len(null_rows) == 0))

    # Test 1.4: Column names (I → Column_I, J kept as junk)
    expected_cols = [
        "Product Info_ID", "Product Info_Name",
        "Q1 Sales_Revenue_Jan", "Q1 Sales_Revenue_Feb", "Q1 Sales_Units_Mar",
        "Q2 Sales_Revenue_Apr", "Q2 Sales_Revenue_May", "Q2 Sales_Units_Jun",
        "Column_I", "asdasd_aaa_a"
    ]
    tests.append(("Column names match", col_names == expected_cols))

    # Test 1.5: First row data (must NOT be eaten by a promote step)
    tests.append(("First row ID = P001", data[0][0] == "P001"))

    # Test 1.6: Last row data (must NOT be eaten by a promote step)
    tests.append(("Last row ID = P006", data[-1][0] == "P006"))

    # Test 1.7: All expected IDs present (two junk rows survive with null ID
    # because they carry values in the junk columns I / J)
    expected_ids = ["P001", "P002", None, "P003", None, "P004", "P005", "P006"]
    actual_ids = [row[0] for row in data]
    tests.append(("All IDs present (incl junk rows)", actual_ids == expected_ids))

    for name, passed in tests:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    if not tests[-2][1]:
        print(f"    Expected: {expected_cols}")
        print(f"    Got:      {col_names}")

    # ========================
    # Test 2: Employee_Data
    # ========================
    print()
    print("=" * 60)
    print("TEST: Employee_Data (3-row merged headers + null rows)")
    print("=" * 60)

    raw2 = read_sheet_raw(filepath, "Employee_Data")
    print(f"  Raw rows: {len(raw2)} (including headers)")

    col_names2, data2 = clean_sheet(raw2, header_rows=3, sep="_")

    print(f"  Column names: {col_names2}")
    print(f"  Data rows: {len(data2)}")

    tests2 = []

    # Test 2.1: Column count
    tests2.append(("Column count = 6", len(col_names2) == 6))

    # Test 2.2: Row count = 4 (6 data rows - 2 null)
    tests2.append(("Row count = 4", len(data2) == 4))

    # Test 2.3: No null rows
    null_rows2 = [i for i, row in enumerate(data2) if is_all_null(row)]
    tests2.append(("No null rows", len(null_rows2) == 0))

    # Test 2.4: Column names
    expected_cols2 = [
        "Emp ID", "Name_First", "Name_Last",
        "Contact_Email", "Contact_Phone", "Contact_Address"
    ]
    tests2.append(("Column names match", col_names2 == expected_cols2))

    # Test 2.5: First row data (must NOT be eaten by a promote step)
    tests2.append(("First row ID = E001", data2[0][0] == "E001"))

    # Test 2.6: All expected IDs present (no data row lost)
    expected_ids2 = ["E001", "E002", "E003", "E004"]
    actual_ids2 = [row[0] for row in data2]
    tests2.append(("All IDs present", actual_ids2 == expected_ids2))

    for name, passed in tests2:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    if not tests2[-2][1]:
        print(f"    Expected: {expected_cols2}")
        print(f"    Got:      {col_names2}")

    # ========================
    # Test 3: Reference (clean sheet, no issues)
    # ========================
    print()
    print("=" * 60)
    print("TEST: Reference (clean sheet, 1-row header, no nulls)")
    print("=" * 60)

    raw3 = read_sheet_raw(filepath, "Reference")
    col_names3, data3 = clean_sheet(raw3, header_rows=1, sep="_")

    print(f"  Column names: {col_names3}")
    print(f"  Data rows: {len(data3)}")

    tests3 = []
    tests3.append(("Column count = 3", len(col_names3) == 3))
    tests3.append(("Row count = 6", len(data3) == 6))
    tests3.append(("No null rows", sum(1 for r in data3 if is_all_null(r)) == 0))
    tests3.append(("Column names match", col_names3 == ["Code", "Description", "Category"]))

    for name, passed in tests3:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    # ========================
    # Test 4: Edge_NullHeaders (some columns null across all 3 header rows)
    # ========================
    print()
    print("=" * 60)
    print("TEST: Edge_NullHeaders (columns with null headers across all rows)")
    print("=" * 60)

    raw4 = read_sheet_raw(filepath, "Edge_NullHeaders")
    print(f"  Raw rows: {len(raw4)} (including headers)")

    col_names4, data4 = clean_sheet(raw4, header_rows=3, sep="_")

    print(f"  Column names: {col_names4}")
    print(f"  Data rows: {len(data4)}")

    tests4 = []

    # Test 4.1: Column count = 4 (all kept, A→Column_A, C→Column_C)
    tests4.append(("Column count = 4", len(col_names4) == 4))

    # Test 4.2: Row count = 3 (4 data rows - 1 null)
    tests4.append(("Row count = 3", len(data4) == 3))

    # Test 4.3: No null rows
    tests4.append(("No null rows", sum(1 for r in data4 if is_all_null(r)) == 0))

    # Test 4.4: No empty column names
    tests4.append(("No empty column names", all(n != "" and n is not None for n in col_names4)))

    # Test 4.5: Expected column names
    # A=null all 3 rows → "Column_A", B="Item", C=null all 3 rows → "Column_C", D="Price"
    expected_cols4 = ["Column_A", "Item", "Column_C", "Price"]
    tests4.append(("Column names match", col_names4 == expected_cols4))

    # Test 4.6: All column names unique
    tests4.append(("All column names unique", len(col_names4) == len(set(col_names4))))

    for name, passed in tests4:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    if not tests4[-2][1]:
        print(f"    Expected: {expected_cols4}")
        print(f"    Got:      {col_names4}")

    # ========================
    # Test 5: Edge_DupHeaders (duplicate header names after combine)
    # ========================
    print()
    print("=" * 60)
    print("TEST: Edge_DupHeaders (duplicate header names after combine)")
    print("=" * 60)

    raw5 = read_sheet_raw(filepath, "Edge_DupHeaders")
    print(f"  Raw rows: {len(raw5)} (including headers)")

    col_names5, data5 = clean_sheet(raw5, header_rows=3, sep="_")

    print(f"  Column names: {col_names5}")
    print(f"  Data rows: {len(data5)}")

    tests5 = []

    # Test 5.1: Column count
    tests5.append(("Column count = 4", len(col_names5) == 4))

    # Test 5.2: Row count = 2 (3 data rows - 1 null)
    tests5.append(("Row count = 2", len(data5) == 2))

    # Test 5.3: No null rows
    tests5.append(("No null rows", sum(1 for r in data5 if is_all_null(r)) == 0))

    # Test 5.4: All column names unique (duplicates suffixed)
    tests5.append(("All column names unique", len(col_names5) == len(set(col_names5))))

    # Test 5.5: Expected column names (ID, Value, Value_2, Value_3)
    expected_cols5 = ["ID", "Value", "Value_2", "Value_3"]
    tests5.append(("Column names match", col_names5 == expected_cols5))

    for name, passed in tests5:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    if not tests5[-1][1]:
        print(f"    Expected: {expected_cols5}")
        print(f"    Got:      {col_names5}")

    # ========================
    # Test 6: Edge_MixedTypes (type detection must sample ALL values, not just the first)
    # ========================
    print()
    print("=" * 60)
    print("TEST: Edge_MixedTypes (numeric column with a stray text value)")
    print("=" * 60)

    raw6 = read_sheet_raw(filepath, "Edge_MixedTypes")
    print(f"  Raw rows: {len(raw6)} (including header)")

    col_names6, data6 = clean_sheet(raw6, header_rows=1, sep="_")
    amount_idx = col_names6.index("Amount")
    amount_values = [row[amount_idx] for row in data6]
    print(f"  Amount values: {amount_values}")

    tests6 = []

    # Test 6.1: Old first-value-only logic misdetects this column as numeric
    old_type = infer_column_type(amount_values, mode="first")
    tests6.append(("Old logic misdetects Amount as number (regression check)", old_type == "number"))

    # Test 6.2: Fixed logic correctly keeps it as text because "N/A" can't convert
    new_type = infer_column_type(amount_values, mode="all")
    tests6.append(("Fixed logic detects Amount as text", new_type == "text"))

    for name, passed in tests6:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    # ========================
    # Test 7: ColumnsToKeep (trim inside the cleaner, before null-row scan)
    # ========================
    print()
    print("=" * 60)
    print("TEST: ColumnsToKeep (early column trim, MissingField.Ignore)")
    print("=" * 60)

    raw7 = read_sheet_raw(filepath, "Sales_Report")
    keep = ["Q1 Sales_Revenue_Jan", "Product Info_ID", "No_Such_Column"]
    col_names7, data7 = clean_sheet(raw7, header_rows=3, sep="_", columns_to_keep=keep)

    print(f"  Kept columns: {col_names7}")
    print(f"  Data rows: {len(data7)}")

    # Reference: full clean (no trim), then trim afterwards — the OLD order.
    full_names, full_data = clean_sheet(raw7, header_rows=3, sep="_")
    ref_names, ref_data = select_columns(full_names, full_data, keep)
    # New order additionally drops rows that are null in every kept column.
    ref_data_dropped = [row for row in ref_data if not is_all_null(row)]

    tests7 = []

    # 7.1: Missing name skipped, order follows the keep list
    tests7.append(("Columns = keep-list order minus missing",
                   col_names7 == ["Q1 Sales_Revenue_Jan", "Product Info_ID"]))

    # 7.2: Same rows as old trim-after order, minus rows null in all kept cols
    tests7.append(("Rows match old order minus all-null-in-kept rows",
                   data7 == ref_data_dropped))

    # 7.3: No row is all-null within the kept columns
    tests7.append(("No all-null rows in kept columns",
                   sum(1 for r in data7 if is_all_null(r)) == 0))

    # 7.4: No columns_to_keep → identical to original behavior
    names_none, data_none = clean_sheet(raw7, header_rows=3, sep="_", columns_to_keep=None)
    tests7.append(("columns_to_keep=None keeps original behavior",
                   names_none == full_names and data_none == full_data))

    for name, passed in tests7:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    # ========================
    # Test 8: Year filter (file-name level, before opening files)
    # ========================
    print()
    print("=" * 60)
    print("TEST: Year filter (ExtractYear + YearFrom/YearTo range)")
    print("=" * 60)

    tests8 = []

    # 8.1: Year extraction from the naming convention
    tests8.append(("Daily_Update_2566.xlsm -> 2566", extract_year("Daily_Update_2566.xlsm") == 2566))
    tests8.append(("name_2556.xlsm -> 2556", extract_year("name_2556.xlsm") == 2556))
    tests8.append(("no-year file -> None", extract_year("summary_final.xlsm") is None))

    # 8.2: No bounds = everything passes
    tests8.append(("No bounds keeps all",
                   all(year_passes_filter(y, None, None) for y in (2560, 2569, None))))

    # 8.3: Range bounds are inclusive
    tests8.append(("2566 in [2566,2567]", year_passes_filter(2566, 2566, 2567)))
    tests8.append(("2567 in [2566,2567]", year_passes_filter(2567, 2566, 2567)))
    tests8.append(("2565 out of [2566,2567]", not year_passes_filter(2565, 2566, 2567)))
    tests8.append(("2568 out of [2566,2567]", not year_passes_filter(2568, 2566, 2567)))

    # 8.4: Single bound works alone
    tests8.append(("YearFrom only: 2569 >= 2566", year_passes_filter(2569, 2566, None)))
    tests8.append(("YearTo only: 2560 <= 2567", year_passes_filter(2560, None, 2567)))

    # 8.5: Null year is KEPT even when a filter is set (never silently dropped)
    tests8.append(("Null year kept under filter", year_passes_filter(None, 2566, 2567)))

    for name, passed in tests8:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    # ========================
    # Test 9: Archive/Current split (CurrentYear mode)
    # ========================
    print()
    print("=" * 60)
    print("TEST: Archive/Current split (CurrentYear = 2569)")
    print("=" * 60)

    cur = 2569
    tests9 = []

    # 9.1: Closed years archive; current year and beyond stay live
    tests9.append(("2566 -> archive", archive_file_passes(2566, cur, None)
                   and not current_file_passes(2566, cur, None)))
    tests9.append(("2569 -> current", current_file_passes(2569, cur, None)
                   and not archive_file_passes(2569, cur, None)))
    tests9.append(("2570 -> current", current_file_passes(2570, cur, None)))

    # 9.2: Null-year files always live, never archived
    tests9.append(("null year -> current, not archive",
                   current_file_passes(None, cur, None)
                   and not archive_file_passes(None, cur, None)))

    # 9.3: YearFrom bounds the archive; YearTo bounds the current side
    tests9.append(("YearFrom=2566 excludes 2565 from archive",
                   not archive_file_passes(2565, cur, 2566)))
    tests9.append(("YearTo=2569 excludes 2570 from current",
                   not current_file_passes(2570, cur, 2569)))

    # 9.4: Every named year lands on EXACTLY one side (no overlap, no gap)
    exactly_one = all(
        archive_file_passes(y, cur, None) != current_file_passes(y, cur, None)
        for y in range(2550, 2581)
    )
    tests9.append(("Named years 2550-2580: exactly one side each", exactly_one))

    for name, passed in tests9:
        status = "PASS" if passed else "FAIL"
        if not passed:
            all_pass = False
        print(f"  [{status}] {name}")

    # ========================
    # Summary
    # ========================
    print()
    print("=" * 60)
    if all_pass:
        print("ALL TESTS PASSED")
    else:
        print("SOME TESTS FAILED — see details above")
    print("=" * 60)

    return all_pass


if __name__ == "__main__":
    success = run_tests()
    exit(0 if success else 1)