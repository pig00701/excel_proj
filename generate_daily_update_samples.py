"""
Generate sample .xlsx files with a "Daily Update" sheet for testing
Daily_Update_Combined.pq / Daily_Update_HeaderList.pq.

Saved as .xlsx (not .xlsm): openpyxl can only write plain-xlsx content —
it cannot embed a real VBA project, so a file it writes with a ".xlsm"
extension has content that doesn't match a true macro-enabled workbook.
Excel then refuses to open it ("file format or file extension is not
valid"). Power Query doesn't care about the extension or macros, so for
local testing you can either use these .xlsx samples directly (temporarily
widen the query's file filter to include .xlsx), or rename a copy to
.xlsm — either way, your real production files (actual macro-enabled
workbooks saved by Excel) are unaffected and will open normally.

Each file's "Daily Update" sheet has:
  Row 1-3: junk/title rows (NOT header data — report title, generated-on
           date, department line)
  Row 4-6: real 3-row merged header (same style as Sales_Report)
  Row 7+ : data

This mirrors the structure Daily_Update_Combined.pq expects: it skips the
first 3 rows, then runs fnCleanMergedHeaders(_, 3, "_") on the rest.
"""
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
title_font = Font(bold=True, size=12)


def build_daily_update_sheet(ws, report_date: str, department: str, rows: list):
    # --- Row 1-3: junk rows, not real headers ---
    ws["A1"] = "Daily Update Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {report_date}"
    ws["A3"] = f"Department: {department}"

    # --- Row 4: top-level category (merged) ---
    ws.merge_cells("A4:B4")
    ws["A4"] = "Order Info"
    ws.merge_cells("C4:E4")
    ws["C4"] = "Metrics"

    # --- Row 5: sub-category ---
    ws["A5"] = "ID"
    ws["B5"] = "Date"
    ws["C5"] = "Qty"
    ws["D5"] = "Price"
    ws["E5"] = "Total"

    # --- Row 6: detail header (same as row 5 here, no further split) ---
    ws["A6"] = "ID"
    ws["B6"] = "Date"
    ws["C6"] = "Qty"
    ws["D6"] = "Price"
    ws["E6"] = "Total"

    for row in range(4, 7):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # --- Data rows starting row 7 ---
    for i, row_data in enumerate(rows):
        row_num = i + 7
        for j, val in enumerate(row_data):
            cell = ws.cell(row=row_num, column=j + 1)
            cell.value = val
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")


FOLDER = r"c:\Users\User\Downloads\new_sb\second-brain\01-Projects\excel_proj"

# Yearly files: the Thai Buddhist year (25xx) at the END of the file name is
# what Daily_Update_Combined.pq's ExtractYear picks up into the "ปี" column.
files = [
    {
        "name": "Daily_Update_2566.xlsx",
        "report_date": "2023-07-01",
        "department": "Sales",
        "rows": [
            ["O001", "2023-07-01", 3, 100, 300],
            ["O002", "2023-07-01", 1, 250, 250],
            [None, None, None, None, None],  # null row
            ["O003", "2023-07-01", 5, 80, 400],
        ],
    },
    {
        "name": "Daily_Update_2567.xlsx",
        "report_date": "2024-07-02",
        "department": "Sales",
        "rows": [
            ["O004", "2024-07-02", 2, 150, 300],
            [None, None, None, None, None],  # null row
            ["O005", "2024-07-02", 4, 90, 360],
        ],
    },
]

for f in files:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Update"
    build_daily_update_sheet(ws, f["report_date"], f["department"], f["rows"])
    output_path = f"{FOLDER}\\{f['name']}"
    wb.save(output_path)
    print(f"Sample saved: {output_path}")
