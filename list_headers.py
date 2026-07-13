"""
list_headers.py — fast replacement for the Daily_Update_HeaderList query.

Power Query's Excel.Workbook must parse the ENTIRE sheet XML of an 80-100 MB
.xlsm before it can hand back even 6 rows, so Daily_Update_HeaderList takes
minutes per run. openpyxl's read_only mode streams rows and stops early, so
this script reads ONLY the junk+header rows and finishes in seconds.

The header-combining logic is imported from tests/test_harness.py — the same
functions that are verified against fnCleanMergedHeaders.pq — so the names
printed here match what Power Query will produce exactly.

Usage:
    python list_headers.py <file-or-folder> [options]

    <file-or-folder>  A single .xlsm/.xlsx file, or a folder — for a folder,
                      the file whose name sorts LAST is used (same "latest
                      year" rule as the M query).

Options (defaults mirror ConfigTable's defaults):
    --sheet NAME      Sheet name to read              (default: Daily Update)
    --junk-rows N     Junk/title rows above header    (default: 3)
    --header-rows N   Merged header rows              (default: 3)
    --separator S     Header level separator          (default: _)
    --first-col L     True Excel letter of the sheet's first used column,
                      same as ConfigTable's FirstColumnLetter (default: A)
    --max-cols N      Column safety cap               (default: 300)
    --csv PATH        Also write the list to a UTF-8 CSV (opens clean in
                      Excel, ready to copy into SelectColumnTable)

Output: one line per column — Excel column letter + cleaned header name.
"""
import argparse
import csv
import sys
import time
from pathlib import Path

import openpyxl

# Reuse the tested header-cleaning logic from the harness
sys.path.insert(0, str(Path(__file__).resolve().parent / "tests"))
from test_harness import (  # noqa: E402
    col_index_to_letter,
    combine_headers,
    fill_down,
    fill_right,
    fix_empty_and_duplicates,
)


def pick_file(target: Path, extensions: tuple[str, ...]) -> Path:
    """A file is used as-is; for a folder, take the name that sorts last
    (latest year, matching the M query), skipping ~$ lock files."""
    if target.is_file():
        return target
    candidates = sorted(
        p for p in target.iterdir()
        if p.suffix.lower() in extensions and not p.name.startswith("~$")
    )
    if not candidates:
        raise SystemExit(f"ไม่พบไฟล์ {'/'.join(extensions)} ใน {target}")
    return candidates[-1]


def read_header_rows(path: Path, sheet: str, n_rows: int, max_cols: int) -> list[list]:
    """Stream only the first n_rows of the sheet — never parses the rest."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise SystemExit(f'ไฟล์ {path.name} ไม่มี sheet ชื่อ "{sheet}" '
                             f"(มี: {', '.join(wb.sheetnames)})")
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=n_rows, values_only=True):
            rows.append(list(row)[:max_cols])
        return rows
    finally:
        wb.close()


def clean_header_names(raw_rows: list[list], junk_rows: int, header_rows: int,
                       sep: str) -> list[str]:
    """Same steps as fnCleanMergedHeaders' header section (steps 1-5)."""
    header_section = [list(r) for r in raw_rows[junk_rows:junk_rows + header_rows]]
    if not header_section:
        raise SystemExit("อ่านแถว header ไม่ได้ — เช็คค่า junk-rows / header-rows")
    # Pad ragged rows to equal width
    width = max(len(r) for r in header_section)
    for r in header_section:
        r.extend([None] * (width - len(r)))

    is_blank_column = [
        all(r[c] is None or str(r[c]).strip() == "" for r in header_section)
        for c in range(width)
    ]
    header_section = [fill_right(r, is_blank_column) for r in header_section]
    header_section = fill_down(header_section)
    return fix_empty_and_duplicates(combine_headers(header_section, sep))


def main() -> None:
    ap = argparse.ArgumentParser(description="List cleaned Daily Update header names (fast).")
    ap.add_argument("target", help="source .xlsm/.xlsx file, or its folder")
    ap.add_argument("--sheet", default="Daily Update")
    ap.add_argument("--junk-rows", type=int, default=3)
    ap.add_argument("--header-rows", type=int, default=3)
    ap.add_argument("--separator", default="_")
    ap.add_argument("--first-col", default="A")
    ap.add_argument("--max-cols", type=int, default=300)
    ap.add_argument("--csv", metavar="PATH", default=None)
    args = ap.parse_args()

    started = time.perf_counter()
    path = pick_file(Path(args.target), (".xlsm", ".xlsx"))
    raw = read_header_rows(path, args.sheet, args.junk_rows + args.header_rows,
                           args.max_cols)
    names = clean_header_names(raw, args.junk_rows, args.header_rows, args.separator)

    # Shift letters so they match the TRUE Excel column (used range may not
    # start at column A — same FirstColumnLetter rule as the M query).
    first = args.first_col.strip().upper()
    offset = 0
    for ch in first:
        offset = offset * 26 + (ord(ch) - 64)
    offset -= 1

    rows = [(col_index_to_letter(i + 1 + offset), name)
            for i, name in enumerate(names)]

    print(f"ไฟล์: {path.name}  |  sheet: {args.sheet}  |  {len(rows)} คอลัมน์"
          f"  |  {time.perf_counter() - started:.1f}s")
    print(f"{'คอลัมน์':<8} ชื่อ header")
    print("-" * 60)
    for letter, name in rows:
        print(f"{letter:<8} {name}")

    if args.csv:
        with open(args.csv, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["ตำแหน่ง column", "ColumnName"])
            w.writerows(rows)
        print(f"\nบันทึก CSV แล้ว: {args.csv} (copy คอลัมน์ ColumnName ไปวางใน SelectColumnTable ได้เลย)")


if __name__ == "__main__":
    main()
